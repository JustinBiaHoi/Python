# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *

# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook("C:\\baitap\\Python\\Lab4_6.xlsx")

# create the sheet object
sheet = wb.active


def excel():
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 10
    sheet.column_dimensions["C"].width = 10
    sheet.column_dimensions["D"].width = 20
    sheet.column_dimensions["E"].width = 20
    sheet.column_dimensions["F"].width = 40
    sheet.column_dimensions["G"].width = 50
    sheet.column_dimensions["H"].width = 50

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "Mã số sinh viên"
    sheet.cell(row=1, column=2).value = "Họ tên"
    sheet.cell(row=1, column=3).value = "Ngày sinh"
    sheet.cell(row=1, column=4).value = "Email"
    sheet.cell(row=1, column=5).value = "Số điện thoại"
    sheet.cell(row=1, column=6).value = "Học kỳ"
    sheet.cell(row=1, column=7).value = "Năm học"
    sheet.cell(row=1, column=8).value = "Chọn môn học"


# Function to set focus (cursor)
def focus1(event):
    # set focus on the course_field box
    mssv_field.focus_set()


# Function to set focus
def focus2(event):
    # set focus on the sem_field box
    Name_field.focus_set()


# Function to set focus
def focus3(event):
    # set focus on the form_no_field box
    NgaySinh_field.focus_set()


# Function to set focus
def focus4(event):
    # set focus on the contact_no_field box
    contact_no_field.focus_set()


# Function to set focus
def focus5(event):
    # set focus on the email_id_field box
    Hocky_field.focus_set()


# Function to set focus
def focus6(event):
    # set focus on the address_field box
    Namhoc_field.focus_set()


# Function for clearing the
# contents of text entry boxes
def clear():
    # clear the content of text entry box
    mssv_field.delete(0, END)
    Name_field.delete(0, END)
    NgaySinh_field.delete(0, END)
    Email_field.delete(0, END)
    contact_no_field.delete(0, END)
    Hocky_field.delete(0, END)
    Namhoc_field.delete(0, END)


# Function to take data from GUI
# window and write to an excel file
def insert():
    # if user not fill any entry
    # then print "empty input"
    if (
        mssv_field.get() == ""
        and Name_field.get() == ""
        and NgaySinh_field.get() == ""
        and Email_field.get() == ""
        and contact_no_field.get() == ""
        and Hocky_field.get() == ""
        and Namhoc_field.get() == ""
    ):
        print("empty input")

    else:
        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column

        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = mssv_field.get()
        sheet.cell(row=current_row + 1, column=2).value = Name_field.get()
        sheet.cell(row=current_row + 1, column=3).value = NgaySinh_field.get()
        sheet.cell(row=current_row + 1, column=4).value = Email_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = Hocky_field.get()
        sheet.cell(row=current_row + 1, column=7).value = Namhoc_field.get()

        # save the file
        wb.save("C:\\baitap\\Python\\Lab4_6.xlsx")

        # set focus on the name_field box
        mssv_field.focus_set()

        # call the clear() function
        clear()


# Driver code
if __name__ == "__main__":
    # create a GUI window
    root = Tk()

    # set the background colour of GUI window
    root.configure(background="light green")

    # set the title of GUI window
    root.title("registration form")

    # set the configuration of GUI window
    root.geometry("500x300")

    excel()

    # create a Form label
    heading = Label(
        root,
        text="THÔNG TIN ĐĂNG KÝ HỌC PHẦN",
        bg="light green",
        fg="red",
        font=("Ariel", 25),
    )

    # create a Name label
    mssv = Label(root, text="MSSV", bg="light green")

    # create a Course label
    name = Label(root, text="Họ tên", bg="light green")

    # create a Semester label
    Ngaysinh = Label(root, text="Ngày sinh", bg="light green")

    # create a Form No. label
    Email = Label(root, text="Email.", bg="light green")

    # create a Contact No. label
    contact_no = Label(root, text="Số điện thoại.", bg="light green")

    # create a Email id label
    Hocky = Label(root, text="Học kỳ", bg="light green")

    # create a address label
    Namhoc = Label(root, text="năm học", bg="light green")

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    heading.grid(row=0, column=1)
    mssv.grid(row=1, column=0)
    name.grid(row=2, column=0)
    Ngaysinh.grid(row=3, column=0)
    Email.grid(row=4, column=0)
    contact_no.grid(row=5, column=0)
    Hocky.grid(row=6, column=0)
    Namhoc.grid(row=7, column=0)

    # create a text entry box
    # for typing the information
    mssv_field = Entry(root)
    Name_field = Entry(root)
    NgaySinh_field = Entry(root)
    Email_field = Entry(root)
    contact_no_field = Entry(root)
    Hocky_field = Entry(root)
    Namhoc_field = Entry(root)

    # bind method of widget is used for
    # the binding the function with the events

    # whenever the enter key is pressed
    # then call the focus1 function
    mssv_field.bind("<Return>", focus1)

    # whenever the enter key is pressed
    # then call the focus2 function
    Name_field.bind("<Return>", focus2)

    # whenever the enter key is pressed
    # then call the focus3 function
    NgaySinh_field.bind("<Return>", focus3)

    # whenever the enter key is pressed
    # then call the focus4 function
    Email_field.bind("<Return>", focus4)

    # whenever the enter key is pressed
    # then call the focus5 function
    contact_no_field.bind("<Return>", focus5)

    # whenever the enter key is pressed
    # then call the focus6 function
    Hocky_field.bind("<Return>", focus6)

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    mssv_field.grid(row=1, column=1, ipadx="100")
    Name_field.grid(row=2, column=1, ipadx="100")
    NgaySinh_field.grid(row=3, column=1, ipadx="100")
    Email_field.grid(row=4, column=1, ipadx="100")
    contact_no_field.grid(row=5, column=1, ipadx="100")
    Hocky_field.grid(row=6, column=1, ipadx="100")
    Namhoc_field.grid(row=7, column=1, ipadx="100")
    Hocky_field.grid(row=8, column=1, ipadx="100")

    # call excel function
    excel()

    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="Black", bg="Red", command=insert)
    submit.grid(row=8, column=1)

    # start the GUI
    root.mainloop()
